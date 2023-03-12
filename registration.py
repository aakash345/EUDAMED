from selenium import webdriver
import pandas as pd
import numpy as np
import openpyxl
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import os
from PyPDF2 import PdfMerger, PdfReader
from io import BytesIO
import json
email = 'hemant@narang.com'
cred = 'Welcome@2023'
url = "https://webgate.ec.europa.eu/eudamed/secure#/actors/view/01FN5XFNVJ8M96HTNA777CB01K"
op = webdriver.ChromeOptions()
# op.add_argument('headless')
path = os.getcwd()
settings = {
       "recentDestinations": [{
            "id": "Save as PDF",
            "origin": "local",
            "account": "",
        }],
        "selectedDestinationId": "Save as PDF",
        "version": 2
    }
prefs = {'printing.print_preview_sticky_settings.appState': json.dumps(settings),
         'savefile.default_directory': path,}
op.add_experimental_option('prefs', prefs)
op.add_argument('--kiosk-printing')
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=op)

driver.get(url)
time.sleep(2)
def login():
    WebDriverWait(driver, 30).until(EC.presence_of_element_located(('name', 'username')))
    email_field = driver.find_element('name','username')
    email_field.send_keys(email)
    button = driver.find_element('name','whoamiSubmit')
    driver.execute_script("arguments[0].click();", button)
    # driver.find_elements_by_css_selector('button.btn-next')[0].click()
    WebDriverWait(driver, 30).until(EC.presence_of_element_located(('name', 'password')))
    pass_field = driver.find_element('name','password')
    pass_field.send_keys(cred)
    button = driver.find_element('name','_submit')
    driver.execute_script("arguments[0].click();", button)

def register_new_page():
    WebDriverWait(driver, 80).until(EC.presence_of_element_located((By.ID, 'nav-menu-expandable-group-click-2')))
    homeLink = driver.find_element(By.ID, 'general-home')
    driver.execute_script("arguments[0].click();", homeLink)
    navLink = driver.find_element(By.ID, 'nav-menu-expandable-group-click-2')
    driver.execute_script("arguments[0].click();", navLink)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "device-device-management-menu-register-new-UDI-DI"))) 
    regisLink = driver.find_element(By.ID, 'device-device-management-menu-register-new-UDI-DI')
    driver.execute_script("arguments[0].click();", regisLink)

def print_submitted(pcode):
    WebDriverWait(driver, 80).until(EC.presence_of_element_located((By.ID, 'nav-menu-expandable-group-click-3')))
    navLink = driver.find_element(By.ID, 'nav-menu-expandable-group-click-3')
    driver.execute_script("arguments[0].click();", navLink)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "priv-device-menu-search-and-view-device"))) 
    searchLink = driver.find_element(By.ID, 'priv-device-menu-search-and-view-device')
    driver.execute_script("arguments[0].click();", searchLink)
    dmodel = driver.find_element(By.ID, "deviceModel")
    driver.execute_script("arguments[0].click();", dmodel)
    dmodel.send_keys(pcode)
    search_btn = driver.find_element(By.ID, "apply-filter")
    driver.execute_script("arguments[0].click();", search_btn)
    WebDriverWait(driver, 80).until(EC.presence_of_element_located((By.ID, 'device-device-management-basic-udi-di-value-column-label')))
    time.sleep(5)
    my_prod = driver.find_element(By.CSS_SELECTOR, "td:nth-child(3) > .truncate")
    driver.execute_script("arguments[0].click();", my_prod)
    WebDriverWait(driver, 80).until(EC.presence_of_element_located((By.ID, 'device-device-management-device-model')))
    result = driver.find_element('id','device-device-management-device-model')
    if(result.text == pcode):
        print("found")
        # Save the page as a PDF
        filename = f'Eudamed - {pcode}.pdf'

        driver.execute_script('window.print();')
        time.sleep(1)
        original_file_name = 'MDR-Eudamed - Search Devices and System or Procedure Packs.pdf' # Replace with the original file name
        os.rename(os.path.join(prefs['savefile.default_directory'], original_file_name),
                os.path.join(prefs['savefile.default_directory'], filename))
    else:
        print_submitted(pcode)


#Page 1
def page1(applicable_regul_val, issuing_entity_val, basic_udi_val):
    applicable_regul_list = ['MDR','IVDR']
    issuing_entity_list = ['GS1','HIBCC','ICCBA','IFA']
    WebDriverWait(driver, 80).until(EC.presence_of_element_located(('id', "basicUdi-code"))) 
    #--->Applicable Regulation
    radiobtn = driver.find_element(By.CSS_SELECTOR, f".radio-button-item:nth-child({applicable_regul_list.index(applicable_regul_val)+1}) span")
    driver.execute_script("arguments[0].click();", radiobtn)
    #---->Issuing Entity
    issuing_entity = driver.find_element(By.CSS_SELECTOR, ".rw-dropdown-list-value")
    driver.execute_script("arguments[0].click();", issuing_entity)
    time.sleep(3)
    dropdown = driver.find_element(By.ID, "basicUdi-issuingAgency_listbox")
    # Find the second option by its text
    option = dropdown.find_element(By.XPATH, f"//div[@class='rw-list-option'][contains(text(),'{issuing_entity_val}')]")
    ActionChains(driver).move_to_element(dropdown).perform()
    driver.execute_script("arguments[0].click();", dropdown)
    WebDriverWait(driver, 90).until(EC.presence_of_element_located((By.XPATH, f"//div[@class='rw-list-option'][contains(text(),'{issuing_entity_val}')]")))
    driver.execute_script("arguments[0].click();", option)
    # opt = driver.find_element(By.CSS_SELECTOR, f".rw-list-option:nth-child({issuing_entity_list.index(issuing_entity_val)+1})")
    # driver.execute_script("arguments[0].click();", opt)

    #---->Basic UDI-DI Code
    basic_udi = driver.find_element(By.ID, "basicUdi-code")
    driver.execute_script("arguments[0].click();", basic_udi)
    basic_udi.clear()
    time.sleep(1)
    basic_udi.send_keys(basic_udi_val)  #Calulator Generated
    #---->System or Procedure
    time.sleep(1)
    systemdev_btn = driver.find_element(By.CSS_SELECTOR, "div:nth-child(1) > .radio-button-group > .radio-button-item:nth-child(1) span")
    driver.execute_script("arguments[0].click();", systemdev_btn)
    #--->Save & Next
    savebtn = driver.find_element(By.CSS_SELECTOR, ".glyphicon-chevron-right")
    driver.execute_script("arguments[0].click();", savebtn)

#Page 2
def page2(risk_class_val, implantable_val, measuring_val, reusable_val, active_val, medicinal_val, pcode, pname):
    options_list = ['Yes','No']
    time.sleep(6)
    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.CLASS_NAME, "rw-dropdown-list-value"))) 
    risk_class = driver.find_element(By.CLASS_NAME, "rw-dropdown-list-value")
    driver.execute_script("arguments[0].click();", risk_class)
    dropdown = driver.find_element(By.ID, "riskClass_listbox")
    # Find the second option by its text
    option = dropdown.find_element(By.XPATH, f"//div[@class='rw-list-option'][contains(text(),'{risk_class_val}')]")
    ActionChains(driver).move_to_element(dropdown).perform()
    dropdown.click()
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, f"//div[@class='rw-list-option'][contains(text(),'{risk_class_val}')]")))
    option.click()

    implantbtn = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(implantable_val)+1}) #implantable")
    driver.execute_script("arguments[0].click();", implantbtn)
        
    measuringfun_btn = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(measuring_val)+1}) #measuringFunction")
    driver.execute_script("arguments[0].click();", measuringfun_btn)
        
    #Update- Working till here (28/02 2:15AM)---------------------------------------------------

    resuable_btn = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(reusable_val)+1}) #reusable")
    driver.execute_script("arguments[0].click();", resuable_btn)
        
    active_btn = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(active_val)+1}) #active")
    driver.execute_script("arguments[0].click();", active_btn)

        
    administering_btn = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(medicinal_val)+1}) #administeringMedicine")
    driver.execute_script("arguments[0].click();", administering_btn)

    WebDriverWait(driver, 50).until(EC.presence_of_element_located(('id', "deviceModel"))) 
    input_box = driver.find_element(By.ID, "deviceModel")
    driver.execute_script("arguments[0].click();", input_box)
    driver.find_element(By.ID, "deviceModel").send_keys(pcode)

    driver.find_element(By.ID, "deviceName").click()
    driver.find_element(By.ID, "deviceName").send_keys(pname)
    save_pg2_btn = driver.find_element(By.ID, "general-button-save-and-next")
    driver.execute_script("arguments[0].click();", save_pg2_btn)
#DONE - SUNDAY

#page3
def page3(issuing_entity_val, emdn_val, trade_name_val, pcode, pname, udi_with_0):
    time.sleep(3)
    WebDriverWait(driver, 50).until(EC.presence_of_element_located(('id', "udiDiData-0--primaryDi-code")))
    issuing_entity_list = ['GS1','HIBCC','ICCBA','IFA']
    dropdn = driver.find_element(By.CSS_SELECTOR, "#udiDiData-0--primaryDi-issuingAgency_input .rw-dropdown-list-value")
    driver.execute_script("arguments[0].click();", dropdn)
    dropopt = driver.find_element(By.CSS_SELECTOR, f".rw-list-option:nth-child({issuing_entity_list.index(issuing_entity_val)+1})")
    driver.execute_script("arguments[0].click();", dropopt)
    slider = driver.find_element(By.CSS_SELECTOR, ".ecl-row:nth-child(3) .slider:nth-child(2)")
    driver.execute_script("arguments[0].click();", slider)
    time.sleep(5)
    emdn = driver.find_element(By.ID, "cndCode")
    driver.execute_script("arguments[0].click();", emdn)
    driver.find_element(By.ID, "cndCode").send_keys(emdn_val)
    find_btn = driver.find_element(By.ID, "general-button-find")
    driver.execute_script("arguments[0].click();", find_btn)
    time.sleep(1)
    trade_input = driver.find_element(By.ID, "udiDiData-0--tradeName-texts-0--text")
    driver.execute_script("arguments[0].click();", trade_input)
    driver.find_element(By.ID, "udiDiData-0--tradeName-texts-0--text").send_keys(trade_name_val)

    #Update 2- Working till here (01/02 4:40AM)---------------------------------------------------

    language1 = driver.find_element(By.CSS_SELECTOR, "#udiDiData-0--tradeName-texts-0--language_input .rw-dropdown-list-value")
    driver.execute_script("arguments[0].click();", language1)
    langopt1 = driver.find_element(By.CSS_SELECTOR, ".rw-list-option:nth-child(7)")
    driver.execute_script("arguments[0].click();", langopt1)

    ref_btn = driver.find_element(By.ID, "udiDiData-0--reference")
    driver.execute_script("arguments[0].click();", ref_btn)
    ref_btn.send_keys(pcode)
    directmarked_btn = driver.find_element(By.ID, "udiDiData-0--directMarking")
    driver.execute_script("arguments[0].click();", directmarked_btn)
    same_udi = driver.find_element(By.CSS_SELECTOR, ".control-label label")
    driver.execute_script("arguments[0].click();", same_udi)
    cb1 = driver.find_element(By.CSS_SELECTOR, ".eu-checkbox:nth-child(1) > .no-class")
    driver.execute_script("arguments[0].click();", cb1)
    cb2 = driver.find_element(By.CSS_SELECTOR, ".eu-checkbox:nth-child(3) > .no-class")
    driver.execute_script("arguments[0].click();", cb2)

    desc = driver.find_element(By.ID, "udiDiData-0--additionalDescription-texts-0--text")
    driver.execute_script("arguments[0].click();", desc)
    desc.send_keys(pname)


    lang_class = driver.find_element(By.NAME, "udiDiData[0].additionalDescription.texts[0].language")
    driver.execute_script("arguments[0].click();", lang_class)
    dropdown = driver.find_element(By.ID, "udiDiData-0--additionalDescription-texts-0--language_listbox")

    # Scroll to the dropdown element to make it clickable
    ActionChains(driver).move_to_element(dropdown).perform()

    # Click on the dropdown element to show the options
    dropdown.click()

    # Wait for the option to be clickable and click on it
    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH,"//div[@id='udiDiData-0--additionalDescription-texts-0--language_listbox']//div[text()='English']")))

    desc_opt = driver.find_element(By.XPATH, "//div[@id='udiDiData-0--additionalDescription-texts-0--language_listbox']//div[text()='English']")
    driver.execute_script("arguments[0].click();", desc_opt)

    udi_stat = driver.find_element(By.CSS_SELECTOR, ".radio-button-item:nth-child(1) span")
    driver.execute_script("arguments[0].click();", udi_stat)
    time.sleep(1)
    udicode = driver.find_element(By.ID, "udiDiData-0--primaryDi-code")
    driver.execute_script("arguments[0].click();", udicode)
    time.sleep(2)
    udicode.send_keys(udi_with_0)
    time.sleep(1)
    page3_btn = driver.find_element(By.ID, "general-button-save-and-next")
    driver.execute_script("arguments[0].click();", page3_btn)


    time.sleep(5)
    elem = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.ID,"udiDiData-0--singleUse")))
    print(elem)
    try:
        # Find the first list item element within the error list element
        text = driver.find_element(By.XPATH, "//ul[contains(., 'Provided value')]/li[1]").text
        # Print the text of the first list item
        print(text)
        def create_new_udi(text, udi_with_0):
            size = len(udi_with_0)
            lis = text.split("should be ")[1][1:2]
            print(lis)
            new_udi = udi_with_0[:-1]+lis
            return new_udi
        new_udi = create_new_udi(text, udi_with_0)
        print(new_udi)
        gotpg3 = driver.find_element(By.CSS_SELECTOR, ".done:nth-child(4) .label-wizard > span")
        driver.execute_script("arguments[0].click();", gotpg3)
        # page 3
        time.sleep(1)
        udicode = driver.find_element(By.ID, "udiDiData-0--primaryDi-code")
        driver.execute_script("arguments[0].click();", udicode)
        time.sleep(2)
        udicode.clear()
        udicode.send_keys(new_udi)
        time.sleep(1)
        page3_btn = driver.find_element(By.ID, "general-button-save-and-next")
        driver.execute_script("arguments[0].click();", page3_btn)
        return new_udi
    except:
        return udi_with_0

#UPDATE 3 - Working till here (02/03 1:26AM)---------------------------------------------------

#page 4
def page4(single_use_val, sterilisation_val, labelled_val, latex_val):
    options_list = ['Yes','No']
    time.sleep(5)
    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.ID,"udiDiData-0--singleUse")))

    pg4_slider1 = driver.find_element(By.CSS_SELECTOR, ".ecl-row:nth-child(2) > .col-md-12:nth-child(1) .slider:nth-child(2)")
    driver.execute_script("arguments[0].click();", pg4_slider1)

    single_use = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(single_use_val)+1}) #udiDiData-0--singleUse")
    driver.execute_script("arguments[0].click();", single_use)
    single_use_slider = driver.find_element(By.ID, "udiDiData-0--maxNumberOfReusesApplicable")
    time.sleep(1)
    driver.execute_script("arguments[0].click();", single_use_slider)

    need_sterlization = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(sterilisation_val)+1}) #udiDiData-0--sterilization")
    driver.execute_script("arguments[0].click();", need_sterlization)

    sterile = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(labelled_val)+1}) #udiDiData-0--sterile")
    driver.execute_script("arguments[0].click();", sterile)

    latex = driver.find_element(By.CSS_SELECTOR, f".eud-u-mr-xs:nth-child({options_list.index(latex_val)+1}) #udiDiData-0--latex")
    driver.execute_script("arguments[0].click();", latex)

    cmr = driver.find_element(By.CSS_SELECTOR, ".eud-u-mr-xs:nth-child(2) #udiDiData-0--cmrSubstance")
    driver.execute_script("arguments[0].click();", cmr)
    endocrine = driver.find_element(By.CSS_SELECTOR, ".eud-u-mr-xs:nth-child(2) #udiDiData-0--endocrineDisruptor")
    driver.execute_script("arguments[0].click();", endocrine)

    time.sleep(2)
    pg4_slider2 = driver.find_element(By.ID, "udiDiData-0--storageApplicable")
    time.sleep(1)
    driver.execute_script("arguments[0].click();", pg4_slider2)

    pg4_slider3 = driver.find_element(By.ID, "udiDiData-0--criticalWarningsApplicable")
    time.sleep(1)
    driver.execute_script("arguments[0].click();", pg4_slider3)

    pg4_btn = driver.find_element(By.ID, "general-button-save-and-next")
    driver.execute_script("arguments[0].click();", pg4_btn)

#page 5
def page5():
    time.sleep(3)
    reprocessed = driver.find_element(By.CSS_SELECTOR, ".eud-u-mr-xs:nth-child(2) #udiDiData-0--reprocessed")
    driver.execute_script("arguments[0].click();", reprocessed)
    intended_purp = driver.find_element(By.CSS_SELECTOR, ".eud-u-mr-xs:nth-child(2) #udiDiData-0--annexXVIApplicable")
    driver.execute_script("arguments[0].click();", intended_purp)
    pg5_slider1 = driver.find_element(By.ID, "udiDiData-0--oemApplicable")
    time.sleep(1)
    driver.execute_script("arguments[0].click();", pg5_slider1)
    pg5_slider2 = driver.find_element(By.ID, "clinicalInvestigationApplicable")
    time.sleep(1)
    driver.execute_script("arguments[0].click();", pg5_slider2)
    human_tissue = driver.find_element(By.CSS_SELECTOR, ".eud-u-mr-xs:nth-child(2) #humanTissues")
    driver.execute_script("arguments[0].click();", human_tissue)
    animal_tissue = driver.find_element(By.CSS_SELECTOR, ".eud-u-mr-xs:nth-child(2) #animalTissues")
    driver.execute_script("arguments[0].click();", animal_tissue)
    medicinal_prod = driver.find_element(By.CSS_SELECTOR, ".eud-u-mr-xs:nth-child(2) #medicinalProduct")
    driver.execute_script("arguments[0].click();", medicinal_prod)
    human_prod = driver.find_element(By.CSS_SELECTOR, ".eud-u-mr-xs:nth-child(2) #humanProduct")
    driver.execute_script("arguments[0].click();", human_prod)
    language1 = driver.find_element(By.CSS_SELECTOR, "#udiDiData-0--placedOnTheMarket_input .rw-dropdown-list-value")
    driver.execute_script("arguments[0].click();", language1)
    langopt1 = driver.find_element(By.CSS_SELECTOR, ".rw-list-option:nth-child(29)")
    driver.execute_script("arguments[0].click();", langopt1)
    pg5_btn = driver.find_element(By.ID, "general-button-save-and-next")
    driver.execute_script("arguments[0].click();", pg5_btn)

def page6():
    pg6_btn = driver.find_element(By.ID, "general-button-submit")
    driver.execute_script("arguments[0].click();", pg6_btn)
    time.sleep(3)
    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.ID,"general-button-submit-my-request")))
    final_btn = driver.find_element(By.ID, "general-button-submit-my-request")
    driver.execute_script("arguments[0].click();", final_btn)
# #dynamic variables
# pcode = 349.070
# pname = "Lane Bone Holding forceps without Ratchet, 300mm"
# udi_without_0 = "890709711279"
# udi_with_0 = "0"+udi_without_0+"0"
# basic_udi_val = "8907097349.0702V"   #calculator generated
# emdn_val = "L091303"
# applicable_regul_val = "MDR"
# issuing_entity_val = "GS1"
# risk_class_val = "Class I"
# implantable_val = "No"
# measuring_val = "No"
# reusable_val = "Yes"
# active_val = "No"
# medicinal_val = "No"
# trade_name_val = "NET"
# single_use_val = "No"
# sterilisation_val = "No"
# labelled_val = "No"
# latex_val = "No"
# df = pd.read_excel('data2.xlsx')
xcel = openpyxl.load_workbook("data2.xlsx")
sh = xcel.active
login()
time.sleep(3)
register_new_page()
time.sleep(3)
# print_submitted("340.010")

for i in range(1, sh.max_row):
    pcode = sh.cell(row = i+1, column = 1).value
    pname = sh.cell(row = i+1, column = 2).value
    udi_without_0 = sh.cell(row = i+1, column = 3).value
    udi_with_0 = "0"+udi_without_0+"0"
    basic_udi_val = sh.cell(row = i+1, column = 4).value   #calculator generated
    emdn_val = sh.cell(row = i+1, column = 6).value
    applicable_regul_val = sh.cell(row = i+1, column = 7).value
    issuing_entity_val = sh.cell(row = i+1, column = 8).value
    risk_class_val = sh.cell(row = i+1, column = 9).value
    implantable_val = sh.cell(row = i+1, column = 10).value
    measuring_val = sh.cell(row = i+1, column = 11).value
    reusable_val = sh.cell(row = i+1, column = 12).value
    active_val = sh.cell(row = i+1, column = 13).value
    medicinal_val = sh.cell(row = i+1, column = 14).value
    trade_name_val = sh.cell(row = i+1, column = 15).value
    single_use_val = sh.cell(row = i+1, column = 16).value
    sterilisation_val = sh.cell(row = i+1, column = 17).value
    labelled_val = sh.cell(row = i+1, column = 18).value
    latex_val = sh.cell(row = i+1, column = 19).value
    status_val = sh.cell(row = i+1, column = 20).value
    try:
        if status_val!='Success':
            page1(applicable_regul_val, issuing_entity_val, basic_udi_val)
            print("page 1 completed")

            page2(risk_class_val, implantable_val, measuring_val, reusable_val, active_val, medicinal_val, pcode, pname)
            print("page 2 completed")

            new_udi = page3(issuing_entity_val, emdn_val, trade_name_val, pcode, pname, udi_with_0)
            sh.cell(row = i+1, column = 5).value = new_udi
            print("page 3 completed")
            
            page4(single_use_val, sterilisation_val, labelled_val, latex_val)
            print("page 4 completed")
            
            page5()
            print("page 5 completed")
            
            page6()
            print("page 6 completed")
            sh.cell(row = i+1, column = 20).value = 'Success'
            time.sleep(3)
            print_submitted(pcode)
            register_new_page()

        else:
            continue
    except:
        sh.cell(row = i+1, column = 20).value = 'Error'
        time.sleep(3)
        register_new_page()
    xcel.save('data2.xlsx')
    time.sleep(2)
xcel.save('data2.xlsx')


# device-registration-link-register-new-basic-udi-di