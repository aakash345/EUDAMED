from selenium import webdriver
import pandas as pd
import numpy as np
import openpyxl
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd

input_gcp = '8907097'

url = "https://www.gs1.org/services/check-character-calculator"
op = webdriver.ChromeOptions()
op.add_argument('headless')
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=op)

xcel = openpyxl.load_workbook('data2.xlsx')
my_sheet_obj = xcel.active
print(my_sheet_obj.max_row)
for i in range(1,my_sheet_obj.max_row):
    pcode = my_sheet_obj.cell(row = i+1, column = 1).value
    driver.get(url)
    name_field = driver.find_element('id','input_gcp')
    name_field.send_keys(input_gcp)
    name2_field = driver.find_element('id','input_model')
    name2_field.send_keys(pcode)

    button = driver.find_element('id','calculate_btn')
    driver.execute_script("arguments[0].click();", button)

    result = driver.find_element('id','ccc-result-digits')
    my_sheet_obj.cell(row = i+1, column = 4).value = result.text
    print(result.text)
xcel.save('data2.xlsx')